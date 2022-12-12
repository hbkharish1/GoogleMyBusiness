
from selenium import webdriver
from openpyxl import load_workbook
import time

from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager

from Google.Google_login import Google
# from posting.posting_login import google_login

# driver = webdriver.Chrome(executable_path=r"/Driver/chromedriver.exe")
# driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
# driver = webdriver.Chrome(executable_path="/Users/mani/Desktop/Python/Driver/chromedriver")
driver = webdriver.Chrome(service=Service("/Users/mani/Desktop/Python/Driver/chromedriver"))
wb = load_workbook(r"/Users/mani/Desktop/Python/Posting/Data/Posting post urls.xlsx")
# wb = load_workbook(r"/posting/excel/posting posts.xlsx")
ws = wb.active

column = ws.max_column
row = ws.max_row

# google_login(driver)
# Google(driver)
google = Google(driver=driver)
google.login()


time.sleep(7)

driver.implicitly_wait(30)
for num in range(420, ws.max_row+1):
    print(num)
    driver.implicitly_wait(30)
    print(ws.cell(row=num, column=2).value)
    driver.get(ws.cell(row=num, column=2).value)
    driver.implicitly_wait(30)
    data = driver.find_elements(By.CLASS_NAME,"p4HiUc")

    for all_post in data:
        for post in all_post.find_elements(By.CLASS_NAME,"LgQiCc"):
            for name in post.find_elements(By.CLASS_NAME,"P9ZBeb"):
                if name.text == "Celebrate with ₹5,000* Cashback. Samsung Galaxy Watch4!":
                    print(name.text)
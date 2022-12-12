from selenium import webdriver
from openpyxl import load_workbook
import time

# from posting.posting_login import google_login
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager

from Google.Google_login import Google

# driver = webdriver.Chrome(executable_path=r"D:\Durai\Driver\chromedriver.exe")
# wb = load_workbook(r"d:\Durai\posting\excel\posting posts.xlsx")
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
wb = load_workbook(r"/Users/mani/Desktop/Python/Posting/Data/Posting post urls.xlsx")
ws = wb.active

column = ws.max_column
row = ws.max_row

google=Google(driver)
google.login()
time.sleep(7)

driver.implicitly_wait(30)
for num in range(2, 433):
    print(num)
    driver.implicitly_wait(30)
    print(ws.cell(row=num, column=2).value)
    driver.get(ws.cell(row=num, column=2).value)
    driver.implicitly_wait(30)
    data = driver.find_elements(By.CLASS_NAME,"p4HiUc")

    for all_post in data:
        for post in all_post.find_elements(By.CLASS_NAME,"LgQiCc"):
            try:
                for name in post.find_elements(By.CLASS_NAME,"P9ZBeb"):
                    if name.text == "Upto ₹5000 Cashback on OnePlus U1S TVs at Poorvika!":
                        print(name.text)
                        post.find_element(By.CLASS_NAME,"JRtysb").click()

            except:
                pass

driver.quit()

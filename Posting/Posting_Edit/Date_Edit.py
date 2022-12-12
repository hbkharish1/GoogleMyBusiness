from selenium import webdriver
from openpyxl import load_workbook
import time
from selenium.webdriver.common.by import By
from GMB.Google.Google_login import Google
from selenium.webdriver.common.keys import Keys
driver = webdriver.Chrome(executable_path=r"D:\Durai\Driver\chromedriver.exe")
wb = load_workbook(r"D:\Durai\GMB\Posting\Data\Posting post urls.xlsx")
ws = wb.active

column = ws.max_column
row = ws.max_row

google = Google(driver=driver)
google.login()
time.sleep(2)

driver.implicitly_wait(10)
for num in range(2, 467):
    print(num)
    driver.implicitly_wait(10)
    print(ws.cell(row=num, column=2).value)
    driver.get(ws.cell(row=num, column=2).value)
    driver.implicitly_wait(10)

    for all_post in driver.find_elements(By.CLASS_NAME,"p4HiUc"):
        for post in all_post.find_elements(By.CLASS_NAME,"LgQiCc"):
            # print(post.text)
            try:
                for name in post.find_elements(By.CLASS_NAME, "P9ZBeb"):
                    if name.text == "Buy Amazon Fire TV Stick Lite & 4K at Lowest Price Today!":
                        print(name.text)
                        post.find_element(By.CLASS_NAME, "JRtysb").click()
                        for post_edit in driver.find_elements(By.CLASS_NAME,"z80M1"):
                            if post_edit.text == "Edit":
                                post_edit.click()
                                time.sleep(3)
                                time.sleep(0.5)
                                driver.find_element(By.ID, 'c12').clear()
                                driver.find_element(By.ID, 'c12').send_keys("""31 May 2022""")
                                driver.implicitly_wait(15)
                                # c81
                                time.sleep(3)
                                for finish1 in driver.find_elements(By.CLASS_NAME, 'bMajjd'):
                                    for finish in driver.find_elements(By.CLASS_NAME, 'VfPpkd-vQzf8d'):
                                        if finish.text == 'Publish':
                                            print('Process completed')
                                            finish.location_once_scrolled_into_view
                                            time.sleep(1.5)
                                            driver.implicitly_wait(30)
                                            finish1.click()
                                            driver.implicitly_wait(30)
                                            time.sleep(4)
                                            print('Published')
            except:
                pass

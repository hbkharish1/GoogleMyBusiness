import time
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
# from Posting.Google.Google_login import Google
from Google.Google_login import Google
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from openpyxl import load_workbook,Workbook
from selenium.common.exceptions import NoSuchElementException
# driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
# driver = webdriver.Chrome(executable_path="/Users/mani/Desktop/Python/Driver/chromedriver")
# driver = webdriver.safari.webdriver.WebDriver(quiet=False)
driver = webdriver.Chrome(service=Service("/Users/mani/Desktop/Python/Driver/chromedriver"))







wb = load_workbook(r"/Users/mani/Desktop/Python/Posting/Data/Posting post urls.xlsx")
ws = wb.active
fb=Workbook()
fs=fb.active


class GooglePosting:

    def create_post(self):
        for post in driver.find_elements(By.CLASS_NAME, 'U26fgb'):  # This loop is for finding the Add Event po up box
            if post.get_attribute("aria-label") == "Create Post":
                post.click()
                time.sleep(1)
                break

    def move_to_event(self):
        time.sleep(2)

        action = ActionChains(driver)
        action.send_keys(Keys.TAB)
        action.send_keys(Keys.ARROW_RIGHT)
        action.send_keys(Keys.ARROW_RIGHT)
        action.send_keys(Keys.ARROW_RIGHT)

        for event in driver.find_elements(By.CLASS_NAME, "v5Gsrd"):
            # print(event.text)
            if event.text == "Product":
                time.sleep(2)
                action.send_keys(Keys.ARROW_RIGHT)
                action.perform()

        for event in driver.find_elements(By.CLASS_NAME, "v5Gsrd"):
            # print(event.text )
            if event.text == "Event":
                event.click()
                break
            else:
                action.perform()

    def image_upload_form(self):
        time.sleep(2)
        for upload_form in driver.find_elements(By.CLASS_NAME, 'BqN5te'):
            if upload_form.text == 'Add photos or videos' or upload_form.text == 'Add photos' :
                upload_form.click()
                time.sleep(1)
                driver.implicitly_wait(30)

    def image_upload(self,**kwargs):
        time.sleep(3)
        test = driver.find_element(By.CLASS_NAME, 'EIlDfe')
        for test1 in test.find_elements(By.CLASS_NAME,'XKSfm-Sx9Kwc'):
            # print('found')
            for test2 in test1.find_elements(By.CLASS_NAME,'XKSfm-Sx9Kwc-bN97Pc'):
                # print('found2')
                for test3 in test2.find_elements(By.CLASS_NAME,'fFW7wc-OEVmcd'):
                    # print('found3')
                    # test9=test3.find_element(By.ID,"ia2vjtwiul92")
                    # driver.switch_to.frame(test3)
                    # print('switched')
                    time.sleep(1)
                    driver.implicitly_wait(10)
                    for test4 in test3.find_elements(By.ID, 'doclist'):
                        # print('found4')
                        # for test5 in test4.find_elements(By.CLASS_NAME,'le-Fc-Sf-Qd-Io'):#<-----------Important---------> fe-Fc-cg-Eo
                        for test5 in test4.find_elements(By.ID, ':c'):#<-----------Important---------> fe-Fc-cg-Eo
                            # print('found5')
                            driver.implicitly_wait(10)
                            for test6 in test5.find_elements(By.ID,':f'):
                                # print('found6')
                                driver.implicitly_wait(10)
                                for test7 in test6.find_elements(By.XPATH,"//input[@type='file']"):
                                # for test7 in test6.find_elements(By.XPATH,"//*[@id=':f']/div"):
                                # for test7 in test6.find_elements(By.CLASS_NAME, "d-u"):
                                    # print('Close the frame now!')
                                    test7.send_keys(kwargs.get('image'))
                                    time.sleep(3)
                                    print("image uploaded")
                                    driver.implicitly_wait(10)
                                    driver.switch_to.default_content()

    def post_title(self,**kwargs):
        time.sleep(0.5)
        try:
            # driver.find_element(By.ID, kwargs.get('id')).clear()
            driver.find_element(By.ID, 'c40').send_keys(kwargs.get('title'))
        except:
            driver.find_element(By.ID, 'c2').send_keys(kwargs.get('title'))
            # driver.find_element(By.ID, kwargs.get('id')).clear()
            # driver.find_element(By.ID, kwargs.get('id')).send_keys(kwargs.get('title'))

    def end_date(self,**kwargs):
        time.sleep(0.5)
        try:
            driver.find_element(By.ID, 'c50').clear()
            driver.find_element(By.ID, 'c50').send_keys(kwargs.get('date'))
        except:
            driver.find_element(By.ID, 'c12').clear()
            driver.find_element(By.ID, 'c12').send_keys(kwargs.get('date'))

    def post_description(self,**kwargs):
        time.sleep(0.5)
        try:
            driver.find_element(By.ID, 'c57').send_keys(kwargs.get('description'))
        except:
            driver.find_element(By.ID, 'c19').send_keys(kwargs.get('description'))

    # def post_title_and_description_date(self,**kwargs):
    #     time.sleep(0.5)
    #     try:
    #         driver.find_element(By.ID, kwargs.get('id')).clear()
    #         driver.find_element(By.ID, kwargs.get('id')).send_keys(kwargs.get('title'))
    #     except:
    #         driver.find_element(By.ID, kwargs.get('id')).clear()
    #         driver.find_element(By.ID, kwargs.get('id')).send_keys(kwargs.get('title'))


    def add_more_click(self):
        time.sleep(0.5)
        for drop_down in driver.find_elements(By.CLASS_NAME, 'VfPpkd-vQzf8d'):
            driver.implicitly_wait(30)
            if drop_down.text == 'Add more details (optional)':
                driver.implicitly_wait(30)
                drop_down.click()

    def none_click(self):
        time.sleep(0.5)
        for drop2 in driver.find_elements(By.CLASS_NAME,'VfPpkd-vQzf8d'):
            if drop2.text == 'None':
                drop2.click()
        driver.implicitly_wait(30)

    def post_call(self,**kwargs):
        time.sleep(0.5)
        for field_call in driver.find_elements(By.CLASS_NAME,'VfPpkd-StrnGf-rymPhb'):
            for field1 in field_call.find_elements(By.CLASS_NAME,"VfPpkd-StrnGf-rymPhb-ibnC6b"):
                driver.implicitly_wait(5)
                if field1.text == kwargs.get('field'):
                    print(kwargs.get('field'))
                    driver.implicitly_wait(30)
                    field1.click()

    def post_published(self):
        time.sleep(0.5)
        for finish1 in driver.find_elements(By.CLASS_NAME, 'bMajjd'):
            for finish in driver.find_elements(By.CLASS_NAME, 'VfPpkd-vQzf8d'):
                if finish.text == 'Publish':
                    print('Process completed')
                    finish.location_once_scrolled_into_view
                    time.sleep(1.5)
                    driver.implicitly_wait(30)
                    finish1.click()
                    driver.implicitly_wait(30)
                    time.sleep(3)
                    print('Published')


class GooglePostingRun:

    def __init__(self, posting=None):
        self.posting = posting

    def range_run(self, **kwargs):
        google = Google(driver=driver)
        google.login()
        time.sleep(15)

        print(self.posting.Title)
        driver.implicitly_wait(3)

        for num in range(kwargs.get('start'), kwargs.get('end')):
            # print(kwargs.get('value'))
            print(num)
            fs.cell(row=1,column=1).value=num
            fb.save("/Users/mani/Desktop/Python/Posting/Posting_lastPost/findlastpost-"+str(kwargs.get('value'))+".xlsx")
            print(ws.cell(row=num, column=2).value)
            driver.get(ws.cell(row=num, column=2).value)
            driver.implicitly_wait(3)
            self.posted()
        driver.close()
        driver.quit()

    def posted(self,**kwargs):
        gp = GooglePosting()
        gp.create_post()
        gp.move_to_event()
        gp.image_upload_form()
        # gp.image_upload(image=self.posting.Img_file)
        gp.image_upload(image=self.posting.Img_file)
        # driver.implicitly_wait(5)
        # time.sleep(3)
        # print(kwargs.get('cell_row'))
        gp.post_title(title=self.posting.Title)
        # gp.post_title_and_description_date(title=self.posting.Title)
        gp.add_more_click()
        gp.end_date(date=self.posting.End_date)
        gp.post_description(description=self.posting.Description)
        # gp.post_title_and_description_date(id='c46', title=self.posting.End_date)
        # gp.post_title_and_description_date(id='c53', title=self.posting.Description)
        gp.none_click()
        gp.post_call(field=self.posting.Field)
        gp.post_published()








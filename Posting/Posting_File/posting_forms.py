import time
import datetime
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
# from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import Select

from Posting.Google.Google_login import Google
from selenium import webdriver
from openpyxl import load_workbook,Workbook
# driver = webdriver.Chrome(service=Service("/Users/mani/Desktop/Python/Driver/chromedriver"))
driver=webdriver.Chrome(service=Service("D:\Harish\Driver\chromedriver_win32\chromedriver.exe"))
date = datetime.date.today().strftime("%#d/%#m/%Y")





wb = load_workbook(r"D:\Harish\Posting\Posting\Data\Posting post urls.xlsx")
ws = wb.active
fb=Workbook()
fs=fb.active


class GooglePosting:

    # def move_to_event(self,**kwargs):
    #     time.sleep(2)
    #     driver.find_element(By.XPATH,'//*[@id="lb"]/div/div[2]/span/undefined/div/div[2]/c-wiz/div/div/div[1]/span[3]/div/button').click()
    #     driver.switch_to.frame(driver.find_element(By.XPATH, value='//*[@id="gsr"]/iframe'))
    #     driver.implicitly_wait(5)
    #     desc=driver.find_element(By.ID,'i4')
    #     desc.send_keys(kwargs.get('title'))
    #     print("switched")
    #     driver.implicitly_wait(1)

    def move_to_frame(self):
        time.sleep(2)
        for button in driver.find_elements(By.CLASS_NAME,'tN4Gf'):
            if button.get_attribute('jsname') == 'z47ex':
                button.click()
        for event in driver.find_elements(By.TAG_NAME,'iframe'):
            if event.get_attribute('sandbox') == 'allow-same-origin allow-scripts allow-forms allow-popups allow-popups-to-escape-sandbox':
                driver.switch_to.frame(event)
                print('Switched')
                driver.implicitly_wait(5)



    def post_title(self,**kwargs):
        time.sleep(0.5)
        desc=driver.find_element(By.ID,'i4')
        desc.send_keys(kwargs.get('title'))
        print("Titled")
        driver.implicitly_wait(1)




    # def image_upload_form(self):
    #     time.sleep(2)
    #     for upload_form in driver.find_elements(By.XPATH, '//*[@id="AH1dze"]/div/div/main/div/div/div/div/div/div[2]/div[1]/div[1]/div/div/div/div/button/span'):
    #         if upload_form.text == 'Add photos or videos' or upload_form.text == 'Add photos' :
    #             upload_form.click()
    #             time.sleep(1)
    #             driver.implicitly_wait(30)

    def image_upload_form(self):
        time.sleep(2)
        for upload in driver.find_elements(By.CLASS_NAME,'hqTBzf'):
            buttonclick=upload.find_element(By.TAG_NAME,'button')
            buttonclick.click()
            time.sleep(1)
            driver.implicitly_wait(30)



    # def image_upload(self, **kwargs):
    #     test3=driver.find_element(By.CLASS_NAME,'picker-dialog-content')
    #     driver.switch_to.frame(test3.find_element(By.TAG_NAME, 'iframe'))
    #     driver.implicitly_wait(5)
    #     print("image")
    #     image=driver.find_element(By.XPATH,'//*[@id="doclist"]/div/div[4]/div[2]/div/div[2]/div/div/div[1]/div/div[2]/input')
    #     # image = driver.find_element(By.TAG_NAME,'input')
    #     image.send_keys(kwargs.get('image'))
    #     print("imageswitched")
    #     time.sleep(3)
    #     driver.implicitly_wait(10)
    #     driver.switch_to.default_content()
    #     driver.switch_to.frame(driver.find_element(By.XPATH, value='//*[@id="gsr"]/iframe'))
    #     driver.implicitly_wait(5)

    def image_upload(self, **kwargs):
        test3=driver.find_element(By.CLASS_NAME,'picker-dialog-content')
        driver.switch_to.frame(test3.find_element(By.TAG_NAME, 'iframe'))
        driver.implicitly_wait(5)
        # print("image")
        image = driver.find_element(By.TAG_NAME,'input')
        image.send_keys(kwargs.get('image'))
        time.sleep(3)
        print("imageuploaded")
        driver.implicitly_wait(10)
        driver.switch_to.default_content()
        for event in driver.find_elements(By.TAG_NAME,'iframe'):
            if event.get_attribute('sandbox') == 'allow-same-origin allow-scripts allow-forms allow-popups allow-popups-to-escape-sandbox':
                driver.switch_to.frame(event)
                driver.implicitly_wait(5)
        driver.implicitly_wait(5)



    def end_date(self,**kwargs):
        time.sleep(0.5)
        # print(date)
        # Start date
        driver.find_element(By.ID,'i7').send_keys(date)
        # End date
        driver.find_element(By.ID,'i14').send_keys(kwargs.get('date'))
        print("Date Updated")


    def post_description(self,**kwargs):
        time.sleep(0.5)
        driver.find_element(By.ID,'i21').send_keys(kwargs.get('description'))





    def post_published(self):
        time.sleep(0.5)
        posts=driver.find_element(By.CLASS_NAME,'FkJOzc')
        for clicked in posts.find_elements(By.TAG_NAME,'button'):
            post = clicked.get_attribute('jsname')
            # print(post)
            if post == "vdQQuc":
                clicked.click()
                print("Successfully Posted")

    def Recapatcha(self):
        robot=driver.find_element(By.ID,'recaptcha')
        driver.switch_to.frame(robot.find_element(By.TAG_NAME,'iframe'))
        driver.find_element(By.CLASS_NAME,'rc-anchor-checkbox').click()
        time.sleep(10)

    def post_call(self,**kwargs):
        # driver.find_element(By.XPATH,'//*[@id="AH1dze"]/div/div/main/div/div/div/div/div/div[1]/div/div[3]/div/div[2]/div/div/div[1]/button/i').click()
        button =driver.find_element(By.CLASS_NAME,'VfPpkd-LgbsSe')
        button.find_element(By.CLASS_NAME,'VfPpkd-kBDsod').click()
        # driver.implicitly_wait(2)
        for field in driver.find_elements(By.CLASS_NAME,'VfPpkd-StrnGf-rymPhb-b9t22c'):
            if field.text == kwargs.get('field'):
                print(kwargs.get('field'))
                field.click()
        # time.sleep(5)



class GooglePostingRun:

    def __init__(self, posting=None):
        self.posting = posting

    def range_run(self, **kwargs):
        google = Google(driver=driver)
        google.login()
        time.sleep(5)

        print(self.posting.Title)
        driver.implicitly_wait(3)
        for num in range(kwargs.get('start'), kwargs.get('end')):
            print(num)
            fs.cell(row=1, column=1).value = num
            if kwargs.get('start') == kwargs.get('end')-1:
                print("changed")
                fs.cell(row=1, column=1).value = kwargs.get("sp")
            fb.save(r"D:\Harish\Posting\Posting\Posting_lastPost\findlastpost-"+str(kwargs.get('value'))+".xlsx")
            print(ws.cell(row=num, column=2).value)
            driver.get(ws.cell(row=num, column=2).value)
            driver.implicitly_wait(3)
            self.posted()
        driver.close()
        driver.quit()



    def posted(self):
        time.sleep(5)
        gp = GooglePosting()
        # gp.Recapatcha()
        gp.move_to_frame()
        gp.post_title(title=self.posting.Title)
        gp.image_upload_form()
        gp.image_upload(image=self.posting.Img_file)
        gp.end_date(date=self.posting.End_date)
        gp.post_description(description=self.posting.Description)
        gp.post_call(field=self.posting.Field)
        # time.sleep(10)
        gp.post_published()








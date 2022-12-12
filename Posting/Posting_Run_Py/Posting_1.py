# from Posting.Posting_File.posting_forms import GooglePostingRun
# from Posting.Posting_Fields.posting_fields import *
from Posting.Posting_File.posting_forms import GooglePostingRun
from Posting.Posting_Fields.posting_fields import *
from openpyxl import load_workbook
wb=load_workbook(r'D:\Harish\Posting\Posting\Posting_lastPost\findlastpost-1.xlsx')
ws=wb.active
g1 = GooglePostingRun(posting=PostingField1)
g1.range_run(start=2, end=101,value=1)
# g1.range_run(start=ws.cell(row=1,column=1).value, end=101,value=1,sp=2)
# 2 to 101
# 71
# g1 = GooglePostingRun(posting=PostingField1)
# g1.range_run(start=2, end=151,value=1)

from Posting_File.posting_forms import GooglePostingRun
from Posting_Fields.posting_fields import *
from openpyxl import load_workbook
wb=load_workbook('/Users/mani/Desktop/Python/Posting/Posting_lastPost/findlastpost-9.xlsx')
ws=wb.active

# 300 to 400
#373

g1 = GooglePostingRun(posting=PostingField2)
# g1.range_run(start=301,end=400,value=9,sp=301)

# g1 = GooglePostingRun(posting=PostingField3)
# g1.range_run(start=416, end=480,value=9)
g1.range_run(start=ws.cell(row=1,column=1).value, end=401,value=9,sp=301)
# g1 = GooglePostingRun(posting=PostingField2)
# g1.range_run(start=334,end=401,value=9)


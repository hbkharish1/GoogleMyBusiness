from Posting_File.posting_forms import GooglePostingRun
from Posting_Fields.posting_fields import *
from openpyxl import load_workbook
wb=load_workbook('/Users/mani/Desktop/Python/Posting/Posting_lastPost/findlastpost-4.xlsx')
ws=wb.active

# 301 to 401
#466 stops
#367
g1 = GooglePostingRun(posting=PostingField1)
# g1.range_run(start=301,end=401,value=4,sp=301)
g1.range_run(start=ws.cell(row=1,column=1).value, end=401,value=4,sp=301)
# g1 = GooglePostingRun(posting=PostingField2)
# g1.range_run(start=2, end=151,value=4)

# g1 = GooglePostingRun(posting=PostingField1)
# g1.range_run(start=444, end=470,value=4)
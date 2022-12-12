# from Posting.Posting_File.posting_forms import GooglePostingRun
# from Posting.Posting_Fields.posting_fields import *
from Posting_File.posting_forms import GooglePostingRun
from Posting_Fields.posting_fields import *
from openpyxl import load_workbook
wb=load_workbook('/Users/mani/Desktop/Python/Posting/Posting_lastPost/findlastpost-2.xlsx')
ws=wb.active
# 101 to 201

# 159
g1 = GooglePostingRun(posting=PostingField1)
# g1.range_run(start=101, end=201,value=2,sp=101)
g1.range_run(start=ws.cell(row=1,column=1).value, end=201,value=2,sp=101)
#283
# g1 = GooglePostingRun(posting=PostingField1)
# g1.range_run(start=284, end=301,value=2)

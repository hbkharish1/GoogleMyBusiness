# from Posting.Posting_File.posting_forms import GooglePostingRun
# from Posting.Posting_Fields.posting_fields import *
from Posting_File.posting_forms import GooglePostingRun
from Posting_Fields.posting_fields import *
from openpyxl import load_workbook
wb=load_workbook('/Users/mani/Desktop/Python/Posting/Posting_lastPost/findlastpost-3.xlsx')
ws=wb.active
# 201 to 301

#261
#466 stops
g1 = GooglePostingRun(posting=PostingField1)
# g1.range_run(start=284, end=301,value=3,sp=201)
g1.range_run(start=ws.cell(row=1,column=1).value, end=301,value=3,sp=201)

# g1 = GooglePostingRun(posting=PostingField1)
# g1.range_run(start=416, end=480,value=3)
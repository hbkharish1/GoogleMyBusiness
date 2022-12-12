from Posting_File.posting_forms import GooglePostingRun
from Posting_Fields.posting_fields import *
from openpyxl import load_workbook
wb=load_workbook('/Users/mani/Desktop/Python/Posting/Posting_lastPost/findlastpost-10.xlsx')
ws=wb.active

# 400 to 465
#466 stops
#408
g1 = GooglePostingRun(posting=PostingField2)
# g1.range_run(start=420,end=480,value=10,sp=401)
g1.range_run(start=ws.cell(row=1,column=1).value, end=480,value=10,sp=401)
# g1 = GooglePostingRun(posting=PostingField2)
# g1.range_run(start=421,end=470,value=10)

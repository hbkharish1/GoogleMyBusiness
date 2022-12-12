from Posting_File.posting_forms import GooglePostingRun
from Posting_Fields.posting_fields import *
from openpyxl import load_workbook
wb=load_workbook('/Users/mani/Desktop/Python/Posting/Posting_lastPost/findlastpost-8.xlsx')
ws=wb.active

# 200 to 300

#74
# g1 = GooglePostingRun(posting=PostingField2)
# g1.range_run(start=361,end=472,value=8)
g1 = GooglePostingRun(posting=PostingField2)
# g1.range_run(start=230,end=301,value=8,sp=201)
# g1 = GooglePostingRun(posting=PostingField3)
# g1.range_run(start=285, end=301,value=8)
g1.range_run(start=ws.cell(row=1,column=1).value, end=301,value=8,sp=201)
# g1 = GooglePostingRun(posting=PostingField2)
# g1.range_run(start=422, end=480,value=8)


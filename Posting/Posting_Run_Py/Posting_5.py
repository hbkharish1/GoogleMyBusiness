from Posting_File.posting_forms import GooglePostingRun
from Posting_Fields.posting_fields import *
from openpyxl import load_workbook
wb=load_workbook('/Users/mani/Desktop/Python/Posting/Posting_lastPost/findlastpost-5.xlsx')
ws=wb.active

# 401 to 465
#471
#408
#415
g1 = GooglePostingRun(posting=PostingField1)
# g1.range_run(start=416,end=480,value=5,sp=401)
g1.range_run(start=ws.cell(row=1,column=1).value, end=480,value=5,sp=401)
# g1 = GooglePostingRun(posting=PostingField2)
# g1.range_run(start=70, end=151,value=5)
# g1 = GooglePostingRun(posting=PostingField2)
# g1.range_run(start=284, end=301,value=5)

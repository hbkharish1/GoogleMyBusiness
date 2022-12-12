from Posting_File.posting_forms import GooglePostingRun
from Posting_Fields.posting_fields import *

from openpyxl import load_workbook
wb=load_workbook('/Users/mani/Desktop/Python/Posting/Posting_lastPost/findlastpost-6.xlsx')
ws=wb.active


# 2 to 100
#466 stops

# 82
g1 = GooglePostingRun(posting=PostingField3)
# g1.range_run(start=135, end=151,value=6)
g1.range_run(start=ws.cell(row=1,column=1).value, end=101,value=11,sp=2)
# g1 = GooglePostingRun(posting=PostingField2)
# g1.range_run(start=416, end=480,value=6)
# g1 = GooglePostingRun(posting=PostingField2)
# g1.range_run(start=151, end=301,value=6)
# g1 = GooglePostingRun(posting=PostingField2)
# g1.range_run(start=45, end=101,value=6,sp=2)
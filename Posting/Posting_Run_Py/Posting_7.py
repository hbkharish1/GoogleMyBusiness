from Posting_File.posting_forms import GooglePostingRun
from Posting_Fields.posting_fields import *
from openpyxl import load_workbook
wb=load_workbook('/Users/mani/Desktop/Python/Posting/Posting_lastPost/findlastpost-7.xlsx')
ws=wb.active


# 78
g1 = GooglePostingRun(posting=PostingField2)
# g1.range_run(start=169,end=201,value=7,sp=101)
# g1 = GooglePostingRun(posting=PostingField2)
# g1.range_run(start=151,end=301,value=7)

# g1 = GooglePostingRun(posting=PostingField3)
# g1.range_run(start=2, end=151,value=7,sp=2)
g1.range_run(start=ws.cell(row=1,column=1).value, end=201,value=7,sp=101)
# g1 = GooglePostingRun(posting=PostingField2)
# g1.range_run(start=301, end=420,value=7)


